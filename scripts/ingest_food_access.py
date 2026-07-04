"""Ingest all USDA Food Access Research Atlas datasets into SQLite.

Handles 4 datasets (2006, 2010, 2015, 2019) with different file formats
(xls, xlsx, csv). Stores:
  - food_access: per-tract food desert flags (per year, for trend analysis)
  - food_access_variables: variable definitions for user-facing documentation

Supports trend analysis: query the same tract across years to see if
food access improved or worsened over time.

Usage:
    python scripts/ingest_food_access.py              # Ingest all available years
    python scripts/ingest_food_access.py --year 2019  # Ingest specific year
    python scripts/ingest_food_access.py --trends       # Show trend analysis for Nashville

Requires:
    pip install openpyxl  # For .xlsx files
    pip install xlrd       # For .xls files (2006 data)
    HUD_API_TOKEN in .env  # For ZIP→tract crosswalk verification
"""
import sys, os, sqlite3, csv, json, time, argparse
# Import food_access_variables directly without triggering the sdoh_profiler package __init__
# (which loads pydantic via shared.models and crashes on Python 3.14)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

# Direct module import (bypasses __init__.py)
import importlib.util

def _load_module(filepath, modname):
    spec = importlib.util.spec_from_file_location(modname, filepath)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[modname] = mod
    spec.loader.exec_module(mod)
    return mod

_fav = _load_module(
    os.path.join(os.path.dirname(__file__), "..", "src", "features", "sdoh_profiler", "providers", "food_access_variables.py"),
    "food_access_variables"
)
FOOD_ACCESS_VARIABLES_2019 = _fav.FOOD_ACCESS_VARIABLES_2019
DATASETS = _fav.DATASETS

# HUD crosswalk is optional (only needed for --trends flag)
# Load it lazily to avoid triggering pydantic import
HUDCrosswalkProvider = None  # Loaded on demand in show_trends()

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
DB_PATH = os.path.join(DATA_DIR, "sdoh.db")

NASHVILLE_ZCTAS = [
    "37115", "37138", "37201", "37203", "37204", "37205", "37206",
    "37207", "37208", "37209", "37210", "37211", "37212", "37213",
    "37214", "37215", "37216", "37217", "37218", "37219", "37220",
    "37221", "37227", "37228", "37229", "37013", "37027", "37076",
]

# Food desert flag columns (present in all versions, though naming varies slightly)
FOOD_DESERT_COLUMNS = {
    2019: {
        "tract": "CensusTract",
        "is_food_desert": "LILATracts_1And10",  # Primary: low-income + low-access at 1/10 mi
        "food_desert_half": "LILATracts_halfAnd10",
        "food_desert_vehicle": "LILATracts_Vehicle",
        "low_income": "LowIncomeTracts",
        "poverty_rate": "PovertyRate",
        "median_income": "MedianFamilyIncome",
        "urban": "Urban",
        "population": "POP2010",
    },
    2015: {
        "tract": "CensusTract",
        "is_food_desert": "LILATracts_1And10",
        "food_desert_half": "LILATracts_halfAnd10",
        "food_desert_vehicle": "LILATracts_Vehicle",
        "low_income": "LowIncomeTracts",
        "poverty_rate": "PovertyRate",
        "median_income": "MedianFamilyIncome",
        "urban": "Urban",
        "population": "POP2010",
    },
    2010: {
        "tract": "CensusTract",
        "is_food_desert": "LILATracts_1And10",
        "food_desert_half": "LILATracts_halfAnd10",
        "low_income": "LowIncomeTracts",
        "poverty_rate": "PovertyRate",
        "median_income": "MedianFamilyIncome",
        "urban": "Urban",
        "population": "POP2010",
    },
    2006: {
        # 2006 Food Desert Locator uses different column names
        "tract": "Tract_FIPS",
        "is_food_desert": None,  # No explicit flag — compute from low income + low access
        "low_income": None,  # Use LOWI column (count, not flag)
        "poverty_rate": "PERCENT_LOWI",  # Percent low income
        "median_income": None,
        "urban": "Urban",
        "population": "TOTALPOP",
        # Custom: 2006 uses count-based, not flag-based
        "low_access_pop": "LOWA_POP",
        "low_access_pct": "PERCENT_LOWA_Pop",
        "low_income_pop": "LOWI",
        "no_vehicle_pop": "HUNV",
        "no_vehicle_pct": "PERCENT_HUNV",
    },
}


def read_csv_dataset(filepath: str) -> list[dict]:
    """Read a CSV file and return list of dicts."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


def read_xlsx_dataset(filepath: str) -> list[dict]:
    """Read an Excel (.xlsx) file and return list of dicts.
    Finds the data sheet (the one with the most rows).
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True)

    # Find the sheet with the most data rows (the data sheet, not metadata)
    best_ws = None
    best_count = 0
    for ws in wb.worksheets:
        row_count = sum(1 for _ in ws.iter_rows())
        if row_count > best_count:
            best_count = row_count
            best_ws = ws

    if best_ws is None or best_count < 10:
        wb.close()
        return []

    rows = list(best_ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val
        result.append(row_dict)
    return result


def read_xls_dataset(filepath: str) -> list[dict]:
    """Read an Excel (.xls) file and return list of dicts.
    Finds the data sheet (the one with the most rows).
    """
    try:
        import xlrd
    except ImportError:
        print("ERROR: xlrd not installed. Run: pip install xlrd")
        sys.exit(1)

    wb = xlrd.open_workbook(filepath)

    # Find the sheet with the most rows
    best_ws = wb.sheet_by_index(0)
    best_rows = best_ws.nrows
    for i in range(wb.nsheets):
        ws = wb.sheet_by_index(i)
        if ws.nrows > best_rows:
            best_rows = ws.nrows
            best_ws = ws

    if best_ws.nrows < 10:
        return []

    headers = [str(best_ws.cell_value(0, c)).strip() for c in range(best_ws.ncols)]
    result = []
    for r in range(1, best_ws.nrows):
        row_dict = {}
        for c in range(best_ws.ncols):
            if c < len(headers):
                row_dict[headers[c]] = best_ws.cell_value(r, c)
        result.append(row_dict)
    return result


def ensure_tables(conn: sqlite3.Connection):
    """Create tables if they don't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS food_access (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tract_geoid TEXT NOT NULL,
            year INTEGER NOT NULL,
            is_food_desert INTEGER,
            food_desert_half_mile INTEGER,
            food_desert_vehicle INTEGER,
            low_income INTEGER,
            poverty_rate REAL,
            median_family_income REAL,
            urban INTEGER,
            population INTEGER,
            source TEXT,
            updated_at TEXT,
            UNIQUE(tract_geoid, year)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS food_access_variables (
            field TEXT PRIMARY KEY,
            long_name TEXT,
            description TEXT,
            year INTEGER
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_food_access_tract ON food_access(tract_geoid)
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_food_access_year ON food_access(year)
    """)
    conn.commit()


def ingest_variables(conn: sqlite3.Connection, year: int):
    """Store variable definitions in the database."""
    if year != 2019:
        return  # Only 2019 has a structured VariableLookup file

    from datetime import datetime
    now = datetime.now().isoformat()

    for field, (long_name, description) in FOOD_ACCESS_VARIABLES_2019.items():
        conn.execute("""
            INSERT OR REPLACE INTO food_access_variables VALUES (?, ?, ?, ?)
        """, (field, long_name, description, year))
    conn.commit()
    print(f"  Stored {len(FOOD_ACCESS_VARIABLES_2019)} variable definitions")


def ingest_year(conn: sqlite3.Connection, year: int, rows: list[dict]):
    """Ingest one year of food access data."""
    from datetime import datetime
    now = datetime.now().isoformat()

    col_map = FOOD_DESERT_COLUMNS.get(year, FOOD_DESERT_COLUMNS[2019])
    tract_col = col_map.get("tract", "CensusTract")

    ingested = 0
    for row in rows:
        tract = str(row.get(tract_col, "")).strip()
        if not tract:
            continue

        # Pad to 11 digits if needed
        if len(tract) < 11:
            tract = tract.zfill(11)

        def safe_int(val) -> int:
            try:
                return int(float(val)) if val is not None else 0
            except (ValueError, TypeError):
                return 0

        def safe_float(val) -> float:
            try:
                return float(val) if val is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        # 2006 doesn't have explicit food desert flag — compute it
        if year == 2006:
            low_access_pct = safe_float(row.get(col_map.get("low_access_pct", ""), 0))
            low_income_pop = safe_int(row.get(col_map.get("low_income_pop", ""), 0))
            total_pop = safe_int(row.get(col_map.get("population", ""), 0))
            # Food desert: significant low-income population AND low access (>33% beyond 1/2 mile)
            is_fd = 1 if (low_access_pct > 33 and low_income_pop > 100) else 0
        else:
            is_fd = safe_int(row.get(col_map.get("is_food_desert", ""), 0))

        fd_half = safe_int(row.get(col_map.get("food_desert_half", ""), 0))
        fd_vehicle = safe_int(row.get(col_map.get("food_desert_vehicle", ""), 0))
        low_income = safe_int(row.get(col_map.get("low_income", ""), 0))
        poverty_rate = safe_float(row.get(col_map.get("poverty_rate", ""), 0))
        median_income = safe_float(row.get(col_map.get("median_income", ""), 0))
        urban = safe_int(row.get(col_map.get("urban", ""), 0))
        population = safe_int(row.get(col_map.get("population", ""), 0))

        conn.execute("""
            INSERT OR REPLACE INTO food_access VALUES (
                NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
        """, (
            tract, year, is_fd, fd_half, fd_vehicle,
            low_income, poverty_rate, median_income, urban, population,
            f"usda_fara_{year}", now,
        ))
        ingested += 1

    conn.commit()
    print(f"  Ingested {ingested:,} tracts for year {year}")
    return ingested


def show_trends(conn: sqlite3.Connection, zip_codes: list[str], hud_token: str):
    """Show food access trends across years for specific ZIP codes."""
    global HUDCrosswalkProvider
    if HUDCrosswalkProvider is None:
        _hud_mod = _load_module(
            os.path.join(os.path.dirname(__file__), "..", "src", "features", "sdoh_profiler", "providers", "hud.py"),
            "hud_provider"
        )
        HUDCrosswalkProvider = _hud_mod.HUDCrosswalkProvider

    hud = HUDCrosswalkProvider(api_token=hud_token)

    print("\n=== Food Access Trend Analysis ===")
    print(f"{'ZIP':<8} {'Year':>6} {'Tracts':>8} {'Food Deserts':>14} {'Poverty Rate':>14} {'Median Income':>14}")
    print("-" * 70)

    for zcta in zip_codes[:10]:  # Limit to 10 for speed
        tracts = hud.get_tracts_for_zip(zcta)
        if not tracts:
            continue

        tract_ids = [t["geoid"] for t in tracts]

        for year in [2006, 2010, 2015, 2019]:
            placeholders = ",".join("?" * len(tract_ids))
            rows = conn.execute(f"""
                SELECT tract_geoid, is_food_desert, poverty_rate, median_family_income
                FROM food_access
                WHERE year = ? AND tract_geoid IN ({placeholders})
            """, [year] + tract_ids).fetchall()

            if rows:
                fd_count = sum(1 for r in rows if r[1])
                avg_poverty = sum(r[2] for r in rows if r[2]) / len(rows)
                avg_income = sum(r[3] for r in rows if r[3]) / len(rows)
                print(f"{zcta:<8} {year:>6} {len(rows):>8} {fd_count:>14} {avg_poverty:>13.1f}% ${avg_income:>12,.0f}")

        time.sleep(0.5)  # HUD rate limit


def main():
    parser = argparse.ArgumentParser(description="Ingest USDA Food Access Atlas data into SQLite")
    parser.add_argument("--year", type=int, choices=[2006, 2010, 2015, 2019],
                        help="Ingest specific year only")
    parser.add_argument("--trends", action="store_true",
                        help="Show food access trend analysis for Nashville ZIPs")
    args = parser.parse_args()

    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    ensure_tables(conn)

    years_to_process = [args.year] if args.year else [2006, 2010, 2015, 2019]

    for year in years_to_process:
        ds = DATASETS.get(year)
        if not ds:
            continue

        filepath = os.path.join(DATA_DIR, ds["folder"], ds["filename"])
        if not os.path.exists(filepath):
            print(f"  {year}: File not found ({filepath}), skipping")
            continue

        print(f"\n=== Processing {ds['name']} ===")
        print(f"  File: {filepath}")

        if ds["format"] == "csv":
            rows = read_csv_dataset(filepath)
        elif ds["format"] == "xlsx":
            rows = read_xlsx_dataset(filepath)
        elif ds["format"] == "xls":
            rows = read_xls_dataset(filepath)
        else:
            print(f"  Unknown format: {ds['format']}")
            continue

        print(f"  Loaded {len(rows):,} rows")

        # Store variable definitions (2019 only has structured lookup)
        ingest_variables(conn, year)

        # Ingest data
        ingest_year(conn, year, rows)

    if args.trends:
        hud_token = os.environ.get("HUD_API_TOKEN", "")
        if not hud_token:
            print("\nERROR: HUD_API_TOKEN not set. Cannot do trend analysis without ZIP→tract crosswalk.")
        else:
            show_trends(conn, NASHVILLE_ZCTAS, hud_token)

    conn.close()

    # Summary
    conn = sqlite3.connect(DB_PATH)
    total = conn.execute("SELECT COUNT(*) FROM food_access").fetchone()[0]
    years = conn.execute("SELECT DISTINCT year FROM food_access ORDER BY year").fetchall()
    vars_count = conn.execute("SELECT COUNT(*) FROM food_access_variables").fetchone()[0]
    conn.close()

    print(f"\n=== Summary ===")
    print(f"Total tract records: {total:,}")
    print(f"Years available: {[y[0] for y in years]}")
    print(f"Variable definitions: {vars_count}")
    print(f"Database: {DB_PATH}")


if __name__ == "__main__":
    main()